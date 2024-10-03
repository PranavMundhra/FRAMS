import cv2
import os
from openpyxl import load_workbook
from datetime import datetime, timedelta

def mark_attendance():
    # Load Excel file
    today = datetime.now()
    month_year = today.strftime("%B_%Y")  # Month and year as sheet name
    excel_file = f"data/attendance_{month_year}.xlsx"

    if not os.path.exists(excel_file):
        print(f"Error: Excel file '{excel_file}' not found.")
        return

    wb = load_workbook(excel_file)
    if month_year not in wb.sheetnames:
        print(f"Error: Sheet '{month_year}' not found in Excel file.")
        return

    ws = wb[month_year]

    # Find the current date column by matching the date string exactly
    date_column = None
    today_str = today.strftime("%d-%m-%Y")  # Ensure consistent date format

    for cell in ws[1]:
        if cell.value == today_str:
            date_column = cell.column
            break

    if date_column is None:
        print(f"Error: Today's date ({today_str}) not found in Excel file.")
        return

    # Load face cascade
    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

    # Start webcam
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        print("Error: Unable to access webcam!")
        return

    end_time = datetime.now() + timedelta(seconds=30)

    # Dictionary to store similarity count for each name
    similarity_count = {}

    while datetime.now() < end_time:
        ret, frame = cap.read()
        if not ret:
            print("Error: Failed to capture image!")
            break

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)

        for (x, y, w, h) in faces:
            # Crop face region
            face_roi = gray[y:y+h, x:x+w]

            # Compare each detected face with the dataset
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
                name = row[0].value
                if name is not None:
                    # Initialize count for the name if not already present
                    similarity_count.setdefault(name, 0)

                    # Specify the directory where the images of each student are stored
                    student_dir = os.path.join("data", name)
                    if os.path.exists(student_dir):
                        # Check if face matches with any student in the dataset
                        for img_name in os.listdir(student_dir):
                            img_path = os.path.join(student_dir, img_name)
                            img = cv2.imread(img_path)
                            img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

                            # Resize stored image to match detected face size
                            img_gray_resized = cv2.resize(img_gray, (face_roi.shape[1], face_roi.shape[0]))

                            # Compare faces using absolute difference
                            diff = cv2.absdiff(face_roi, img_gray_resized)
                            similarity = diff.mean()

                            # Increment count if similarity threshold is met
                            if similarity < 35:  # Lower value means more similar
                                similarity_count[name] += 1
                                print(f"Match found for {name}: {similarity}")

        cv2.imshow('Webcam', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

    # Save attendance
    try:
        for name_inner, count in similarity_count.items():
            # Find the cell in the row for the name and current date
            for row_inner in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
                if row_inner[0].value == name_inner:
                    # Get the corresponding cell in the current date column
                    cell = ws.cell(row=row_inner[0].row, column=date_column)

                    if cell.value is None:
                        if count >= 25:
                            cell.value = "P"  # Mark present
                            print(f"Attendance marked as Present for {name_inner} on {today_str}")
                        else:
                            cell.value = "A"  # Mark absent
                            print(f"Attendance marked as Absent for {name_inner} on {today_str}")
                    else:
                        print(f"Attendance already marked for {name_inner} on {today_str}")
                    break

        wb.save(excel_file)
        print("Attendance saved.")
    except Exception as e:
        print(f"Error saving attendance: {e}")

# Function to append names at the bottom
def append_names(ws):
    student_dirs = os.listdir("data")

    # Start appending student names from the second row
    next_row = 2  # Start from A2 to keep A1 empty
    for student_folder in student_dirs:
        if not any(row[0].value == student_folder for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1)):
            while ws.cell(row=next_row, column=1).value is not None:  # Find the first empty row
                next_row += 1
            ws.cell(row=next_row, column=1, value=student_folder)

    return ws

if __name__ == "__main__":
    mark_attendance()
