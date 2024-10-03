import streamlit as st
import cv2
import os
from openpyxl import load_workbook, Workbook
from datetime import datetime
import subprocess
def collect_data():
    st.title("Facial Recognition Attendance System")

    student_name = st.text_input("Enter student name:")
    
    if st.button("Start Collecting Data"):
        if not student_name:
            st.error("Please enter student name.")
        else:
            collect_face_data(student_name)

def collect_face_data(student_name):
    # Ensure the Excel file and student names are updated before data collection
    update_excel()

    # Now proceed with collecting face data
    student_dir = os.path.join("data", student_name)
    if not os.path.exists(student_dir):
        os.makedirs(student_dir)

    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        st.error("Unable to access webcam!")
        return

    count = 0
    while count < 150:
        ret, frame = cap.read()
        if not ret:
            st.error("Failed to capture image!")
            break

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml').detectMultiScale(gray, 1.3, 5)

        for (x, y, w, h) in faces:
            cv2.imwrite(f"{student_dir}/face_{count}.jpg", frame[y:y+h, x:x+w])
            count += 1

        if count == 50:
            st.info("Rotate your face slightly to the right.")
        if count == 100:
            st.info("Rotate your face slightly to the left.")

    cap.release()
    cv2.destroyAllWindows()

    if count < 150:
        st.warning(f"Only {count} images were captured for {student_name}. Please retry.")
    else:
        st.success(f"Data for {student_name} added successfully!")

    cap.release()
    cv2.destroyAllWindows()

    if count < 150:
        st.warning(f"Only {count} images were captured for {student_name}. Please retry.")
    else:
        update_excel(student_name)
        st.success(f"Data for {student_name} added successfully!")

        
def update_excel():
    today = datetime.now()
    month_year = today.strftime("%B_%Y")
    excel_file = f"data/attendance_{month_year}.xlsx"

    # Check if the Excel file already exists, if not create and initialize it
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = month_year

        # Set the header for names in A1
        ws.cell(row=1, column=1, value="Names")

        # Add the dates of the current month to the first row (columns 2 onward)
        for day in range(1, 32):
            try:
                date = today.replace(day=day)
                ws.cell(row=1, column=day + 1, value=date.strftime("%d-%m-%Y"))
            except ValueError:
                break  # Handles shorter months

        wb.save(excel_file)
    else:
        wb = load_workbook(excel_file)

    # Ensure the correct sheet is being used
    if month_year not in wb.sheetnames:
        ws = wb.create_sheet(month_year)
        ws.cell(row=1, column=1, value="Names")
        for day in range(1, 32):
            try:
                date = today.replace(day=day)
                ws.cell(row=1, column=day + 1, value=date.strftime("%d-%m-%Y"))
            except ValueError:
                break
    else:
        ws = wb[month_year]

    # Add student names from the 'data' directory if not already present
    student_dirs = os.listdir("data")
    row = 2  # Start adding student names from row 2
    for student_folder in student_dirs:
        if not any(row[0].value == student_folder for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1)):
            ws.cell(row=row, column=1, value=student_folder)
            row += 1

    wb.save(excel_file)

def add_dates(ws, today):
    """This function ensures all dates of the month are added to the first row."""
    last_column = ws.max_column
    
    # Check if the first row already contains date headers
    existing_dates = [ws.cell(row=1, column=col).value for col in range(2, last_column + 1)]
    
    for day in range(1, 32):
        try:
            date = today.replace(day=day)
            date_str = date.strftime("%d-%m-%Y")
            
            if date_str not in existing_dates:
                ws.cell(row=1, column=last_column + 1, value=date_str)
                last_column += 1
        except ValueError:
            break  # Handles shorter months

# Call the update_excel function
update_excel()

if __name__ == "__main__":
    collect_data()
