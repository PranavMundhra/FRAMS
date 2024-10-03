import streamlit as st
import cv2
import os
from openpyxl import load_workbook, Workbook
from datetime import datetime

def update_excel():
    today = datetime.now()
    month_year = today.strftime("%B_%Y")
    excel_file = f"data/attendance_{month_year}.xlsx"

    # Check if the Excel file already exists, if not create and initialize it
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = month_year

        # Add the dates of the current month to the first row (columns 2 onward)
        ws.cell(row=1, column=1, value="Names")
        add_dates(ws, today)
        
        wb.save(excel_file)
    else:
        wb = load_workbook(excel_file)

    # Ensure the correct sheet is being used
    if month_year not in wb.sheetnames:
        ws = wb.create_sheet(month_year)
        ws.cell(row=1, column=1, value="Names")
        add_dates(ws, today)
    else:
        ws = wb[month_year]
        
        # Ensure all dates are present even in existing sheets
        add_dates(ws, today)

    # Add student names from the 'data' directory if not already present
    student_dirs = [folder for folder in os.listdir("data") if os.path.isdir(os.path.join("data", folder))]
    
    row = 2
    for student_folder in student_dirs:
        if not any(row[0].value == student_folder for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1)):
            ws.cell(row=row, column=1, value=student_folder)
            row += 1

    wb.save(excel_file)

def add_dates(ws, today):
    """This function ensures all dates of the month are added to the first row."""
    last_column = ws.max_column
    
    # Check if the first row already contains date headers
    existing_dates = [ws.cell(row=1, column=col).value for col in range(2, last_column + 1)]
    
    for day in range(1, 32):
        try:
            date = today.replace(day=day)
            date_str = date.strftime("%d-%m-%Y")
            
            if date_str not in existing_dates:
                ws.cell(row=1, column=last_column + 1, value=date_str)
                last_column += 1
        except ValueError:
            break  # Handles shorter months

# Call the update_excel function
update_excel()
